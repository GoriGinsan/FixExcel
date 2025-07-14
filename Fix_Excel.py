import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import win32com.client as client
import threading
import time
from PIL import Image, ImageGrab, ImageFile
import tempfile
import ctypes.wintypes
import openpyxl

supported_formats = {'.xls': 56, '.xlsx': 51, '.xlsm': 52, '.xlsb': 50}

class ExcelLightenApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 軽量化・最適化ツール")
        self.geometry("710x420")
        self.configure(bg="#012340")
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.gif_window = None

        # ヘッダー
        header = tk.Frame(self, bg="#027333", height=60)
        header.pack(fill="x", pady=(16, 10))
        tk.Label(header, text="Excel修理キット", bg="#027333", fg="white",
                 font=("メイリオ", 20, 'bold')).pack(pady=(7,2))
        tk.Label(header, text="（DRMのままでも動きます）", bg="#027333", fg="white",
                 font=("メイリオ", 12)).pack()

        # ファイル選択
        frame_path = tk.Frame(self, bg="#012340")
        frame_path.pack(fill="x", padx=30, pady=(8, 16))
        frame_path.columnconfigure(1, weight=1)
        tk.Label(frame_path, text="ファイルパス：", bg="#012340", fg="white",
                 font=("メイリオ", 13)).grid(row=0, column=0, sticky='e', padx=(0,10))
        self.path_var = tk.StringVar()
        tk.Entry(frame_path, textvariable=self.path_var, width=46,
                 font=("メイリオ", 13)).grid(row=0, column=1, padx=(0,10), ipady=2)
        tk.Button(frame_path, text="参照", width=9, bg="#03A63C", fg="white",
                  font=("メイリオ",13,"bold"), command=self.browse_file
        ).grid(row=0, column=2, padx=(0,3), ipady=2)

        # 数式→値 オプションのみ残す
        optf = tk.Frame(self, bg="#012340")
        optf.pack(anchor="w", padx=44, pady=(0, 6))
        self.formula_to_value = tk.BooleanVar(value=True)
        tk.Checkbutton(optf, text="数式を値に変換", variable=self.formula_to_value,
            bg="#012340", fg="white", font=("メイリオ",12), selectcolor="#027333").pack(side="left", padx=10)

        # ステータスと進行度
        status_frame = tk.Frame(self, bg="#012340")
        status_frame.pack(fill="x", padx=40, pady=(10,6))
        self.status_var = tk.StringVar(value="実行ボタンを押してください")
        self.progress_var = tk.DoubleVar(value=0)
        status_label = tk.Label(
            status_frame, textvariable=self.status_var,
            fg="white", bg="#012340", font=("メイリオ", 12, "bold")
        )
        status_label.pack(anchor="w", pady=(0,7), padx=3)
        self.progress = ttk.Progressbar(
            status_frame, variable=self.progress_var,
            maximum=100, length=540, mode="determinate"
        )
        self.progress.pack(pady=(0,2), padx=(0,12))

        btn_frame = tk.Frame(self, bg="#012340")
        btn_frame.pack(pady=(18, 8))
        self.run_btn = tk.Button(btn_frame, text="実行", width=12, height=2,
                                 bg="#04D939", fg="white",
                                 font=("メイリオ",16,"bold"),
                                 command=self.threaded_lighten_excel_file)
        self.run_btn.pack(pady=(0,4))

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="Excelファイルを選択",
            filetypes=[("Excel files", "*.xls;*.xlsx;*.xlsm;*.xlsb")]
        )
        if path:
            self.path_var.set(path)

    def get_desktop_path(self):
        CSIDL_DESKTOP = 0
        SHGFP_TYPE_CURRENT = 0
        buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
        ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOP, None, SHGFP_TYPE_CURRENT, buf)
        return buf.value

    def set_progress(self, val, msg=None):
        self.progress_var.set(val)
        if msg is not None:
            self.status_var.set(msg)
        self.update_idletasks()

    def threaded_lighten_excel_file(self):
        self.run_btn.config(state="disabled")
        self.show_gif()
        t = threading.Thread(target=self.lighten_excel_file)
        t.start()

    def show_gif(self):
        if self.gif_window:
            return
        gif_path = os.path.join(os.path.dirname(__file__), "jojo.gif")
        try:
            from PIL import Image, ImageTk, ImageSequence
            gif = Image.open(gif_path)
            frames = [ImageTk.PhotoImage(frame.copy().convert('RGBA')) for frame in ImageSequence.Iterator(gif)]
        except Exception as e:
            return  # GIFが無ければ何も表示しない

        self.gif_window = tk.Toplevel(self)
        self.gif_window.overrideredirect(True)
        self.gif_window.attributes('-topmost', True)
        self.gif_window.attributes('-alpha', 0.85)

        # 画面右下に配置
        user32 = ctypes.windll.user32
        screen_width = user32.GetSystemMetrics(0)
        screen_height = user32.GetSystemMetrics(1)
        x = screen_width - 210
        y = screen_height - 170
        self.gif_window.geometry(f"180x127+{x}+{y}")
        label = tk.Label(self.gif_window, bg="#012340")
        label.pack()

        def animate(index=0):
            frame = frames[index]
            label.configure(image=frame)
            self.gif_window.after(80, animate, (index + 1) % len(frames))
        animate()

    def close_gif(self):
        if self.gif_window:
            self.gif_window.destroy()
            self.gif_window = None

    def lighten_excel_file(self):
        try:
            file_path = self.path_var.get().strip().strip('"')
            if not file_path or not os.path.isfile(file_path):
                messagebox.showerror("エラー", "有効なExcelファイルを指定してください。")
                self.run_btn.config(state="normal")
                self.close_gif()
                return

            self.set_progress(0, "Excel処理開始…")
            base, ext = os.path.splitext(file_path)
            ext = ext.lower()
            file_format = supported_formats.get(ext, 51)

            # デスクトップに保存
            desktop_path = self.get_desktop_path()
            filename = os.path.basename(base)
            cleaned_path = os.path.join(desktop_path, f"{filename}_light{ext}")

            if os.path.exists(cleaned_path):
                try:
                    os.remove(cleaned_path)
                except Exception as e:
                    messagebox.showerror("エラー", f"コピー先ファイルの削除に失敗しました: {e}")
                    self.run_btn.config(state="normal")
                    self.close_gif()
                    return

            excel = client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            try:
                wb = excel.Workbooks.Open(file_path, UpdateLinks=0, ReadOnly=False)
            except Exception as e:
                self.set_progress(0, "ワークブックを開けませんでした。")
                messagebox.showerror("エラー", f"ワークブックを開けませんでした: {e}")
                excel.Quit()
                self.run_btn.config(state="normal")
                self.close_gif()
                return

            try:
                # 1. 外部リンク削除
                self.set_progress(5, "外部リンク削除中…")
                try:
                    links = wb.LinkSources(1)
                except:
                    links = None
                n_cut = 0
                if links:
                    for link in links:
                        try:
                            wb.BreakLink(Name=link, Type=1)
                            n_cut += 1
                        except:
                            pass

                # 2. 名前定義クリーニング（必ず実行）
                self.set_progress(10, "名前定義クリーニング…")
                n_delname = 0
                try:
                    names = list(wb.Names)
                    for name in reversed(names):
                        try:
                            name.Delete()
                            n_delname += 1
                        except:
                            pass
                except:
                    pass

                # 3. ワークシートごとに最適化
                n_ws, n_clearfmt = 0, 0
                n_sheets = wb.Worksheets.Count
                for i, ws in enumerate(wb.Worksheets):
                    self.set_progress(10 + (70*i/n_sheets), f"シート{n_ws+1}/{n_sheets}最適化…")

                    # 画像圧縮（Shapeがあれば自動処理）
                    try:
                        self.compress_all_shapes(ws)
                    except Exception:
                        pass

                    # 数式→値
                    if self.formula_to_value.get():
                        try:
                            rng = ws.UsedRange
                            rng.Value = rng.Value
                        except Exception:
                            try:
                                rng.Copy()
                                paste_val = getattr(client.constants, 'xlPasteValues', -4163)
                                rng.PasteSpecial(Paste=paste_val)
                                excel.CutCopyMode = False
                            except:
                                pass

                    # ゴーストデータ消し
                    try:
                        last_cell = ws.UsedRange.Cells(ws.UsedRange.Rows.Count, ws.UsedRange.Columns.Count)
                        max_row = ws.Cells.SpecialCells(11).Row
                        max_col = ws.Cells.SpecialCells(11).Column
                        if max_row > ws.UsedRange.Rows.Count:
                            ws.Rows(f"{ws.UsedRange.Rows.Count+1}:{max_row}").Delete()
                        if max_col > ws.UsedRange.Columns.Count:
                            ws.Columns(f"{ws.UsedRange.Columns.Count+1}:{max_col}").Delete()
                    except:
                        pass

                    # 未使用セル書式クリア
                    try:
                        maxrow = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1
                        maxcol = ws.UsedRange.Column + ws.UsedRange.Columns.Count - 1
                        nrows = ws.Rows.Count
                        ncols = ws.Columns.Count
                        if maxrow < nrows:
                            rg = ws.Range(ws.Cells(maxrow+1,1), ws.Cells(nrows, ncols))
                            rg.ClearFormats()
                            n_clearfmt += rg.Count
                        if maxcol < ncols:
                            rg = ws.Range(ws.Cells(1,maxcol+1), ws.Cells(maxrow, ncols))
                            rg.ClearFormats()
                            n_clearfmt += rg.Count
                    except:
                        pass

                    # コメント・ハイパーリンク全削除
                    try:
                        for c in ws.Comments: c.Delete()
                    except:
                        pass
                    try:
                        ws.Hyperlinks.Delete()
                    except:
                        pass

                    n_ws += 1

                self.set_progress(86, "カスタムスタイルクリーニング…")
                try:
                    styles = wb.Styles
                    if styles.Count > 256:
                        for i in range(styles.Count, 257, -1):
                            try:
                                styles.Item(i).Delete()
                            except:
                                pass
                except:
                    pass

                self.set_progress(97, "保存中…")
                try:
                    wb.SaveAs(cleaned_path, FileFormat=file_format)
                except Exception as e_save:
                    cleaned_xlsx = os.path.join(desktop_path, f"{filename}_fix.xlsx")
                    wb.SaveAs(cleaned_xlsx, FileFormat=51)
                    cleaned_path = cleaned_xlsx

                self.close_gif()
                self.set_progress(100, "完了！")
                messagebox.showinfo(
                    "処理完了",
                    f"リンク{n_cut}件, 名前{n_delname}個削除, 書式クリア{n_clearfmt}セル, シート{n_ws}枚最適化\n\n"
                    f"保存先:\n{cleaned_path}"
                )
            except Exception as e_all:
                # 失敗時リカバリー
                self.set_progress(0, "エラー発生。復元処理…")
                self.close_gif()
                result = self.salvage_excel(file_path)
                if result:
                    messagebox.showinfo(
                        "復元完了", f"Excel復元に失敗しました。\n\nセル値だけ復元したxlsxを保存しました：\n{result}"
                    )
                else:
                    messagebox.showerror("復元不可", f"復元も失敗：{e_all}")

        finally:
            try:
                wb.Close(SaveChanges=False)
            except:
                pass
            try:
                excel.Quit()
            except:
                pass
            self.run_btn.config(state="normal")
            self.close_gif()

    # --- 画像圧縮 ---
    def compress_all_shapes(self, ws):
        shape_count = ws.Shapes.Count
        if shape_count == 0:
            return
        for shape in ws.Shapes:
            try:
                # msoPicture=13, msoLinkedPicture=11
                if shape.Type in [13, 11]:
                    l, t, w, h = shape.Left, shape.Top, shape.Width, shape.Height
                    shape.Copy()
                    # 画像をクリップボードから取得
                    time.sleep(0.2)
                    img = ImageGrab.grabclipboard()
                    if img:
                        temp_path = os.path.join(tempfile.gettempdir(), f"temp_{time.time()}.png")
                        img.save(temp_path)
                        # Pillowで再圧縮
                        ImageFile.LOAD_TRUNCATED_IMAGES = True
                        img2 = Image.open(temp_path)
                        img2 = img2.convert("RGB")
                        img2.save(temp_path, optimize=True, quality=65)
                        # 元画像削除
                        shape.Delete()
                        ws.Pictures().Insert(temp_path)
                        # 最後のShapeを再取得
                        new_shape = ws.Shapes(ws.Shapes.Count)
                        new_shape.Left = l
                        new_shape.Top = t
                        new_shape.Width = w
                        new_shape.Height = h
                        os.remove(temp_path)
            except Exception as e:
                continue

    # --- 失敗時、値だけ復元 ---
    def salvage_excel(self, file_path):
        try:
            # xls/xlsmはサポート外
            if not file_path.lower().endswith(".xlsx"):
                return None
            wb_src = openpyxl.load_workbook(file_path, data_only=True)
            wb_new = openpyxl.Workbook()
            wb_new.remove(wb_new.active)
            for ws_src in wb_src.worksheets:
                ws_new = wb_new.create_sheet(ws_src.title)
                for r, row in enumerate(ws_src.iter_rows(), 1):
                    for c, cell in enumerate(row, 1):
                        ws_new.cell(r, c).value = cell.value
            out_path = os.path.join(self.get_desktop_path(), "salvage_" + os.path.basename(file_path))
            wb_new.save(out_path)
            return out_path
        except Exception as e:
            return None

    def on_close(self):
        self.destroy()

if __name__ == "__main__":
    app = ExcelLightenApp()
    app.mainloop()
